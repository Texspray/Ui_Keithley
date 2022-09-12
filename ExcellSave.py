from openpyxl import Workbook
from openpyxl.utils import  get_column_letter
#from Janela2 import CustomDialog

GateDreno='g'
def SalvarDocumento(NomeArquivo, listaA, listaB, listaC, listaD):
    wb = Workbook()
    dest_filename = "{}.xlsx".format(NomeArquivo)
    ws1 = wb.active
    ws1.title = "Medidas"
    ws1['A1'] = 'V_varredura'
    ws1['B1'] = 'Corrente'
    ws1['C1'] = 'Voltagem_Stepper'
    ws1['D1'] = 'Corrente Gate'
    print(listaA)
    print(listaB)
    print(listaC)
    print(listaD)
    row_number = 2
    your_column = 1
    impar = 1
    print("O tamanho da lista B é: ", len(listaB))
    for i, value in enumerate(listaA[0], start=row_number):
        ws1.cell(row=i, column=your_column).value = value
    for n in range(len(listaB)):
        ws1.cell(row=1, column=your_column+1+2*n).value = 'id{0} V{2} ={1}'.format(n, listaC[n], GateDreno)
        ws1.cell(row=1, column=2+impar).value = 'ig{0} V{2} ={1}'.format(n, listaC[n], GateDreno)
        for i, value in enumerate(listaB[n], start=row_number):
            ws1.cell(row=i, column=your_column+1+2*n).value = value
        for i, value in enumerate(listaD[n], start=row_number):
            ws1.cell(row=i, column=2+impar).value = value

        """for i, value in enumerate(listaC[n], start=row_number):
            ws1.cell(row=i, column=your_column+2*len(listaB)+n+1).value = value"""
        impar = impar + 2
    wb.save(filename= dest_filename)

def salvarTempo(fileName, id, ig):
    wb = Workbook()
    ws1 = wb.active
    dest_filename = '{}.xlsx'.format(fileName)
    ws1.title = "Medidas"
    ws1['A1'] = 'Corrente Canal A'
    ws1['B1'] = 'Corrente Canal B'
    row_number = 2
    your_column = 1
    for i, value in enumerate(id, start=row_number):
        ws1.cell(row=i, column=your_column).value = value
    for i, value in enumerate(ig, start=row_number):
        ws1.cell(row=i, column=your_column + 1).value = value
    wb.save(filename=dest_filename)

    #CustomDialog.close()
